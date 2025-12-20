"""
Monitoring Page Views - Read-only monitoring with time filters and export
"""

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.utils import timezone
from datetime import datetime, timedelta
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from . import models

# Import existing configurations from views.py
from .views import (
    MACHINE_CONFIGS, 
    ASSEMBLY_CONFIGS, 
    OP80_CONFIG,
    parse_timestamp_to_datetime
)


def get_all_model_names_monitoring():
    """Get distinct model names from all machines"""
    model_names = set()
    
    # Get from CNC machines
    for config in MACHINE_CONFIGS:
        if config['prep_model'] and hasattr(config['prep_model'], 'objects'):
            try:
                prep_models = config['prep_model'].objects.values_list('model_name', flat=True).distinct()
                model_names.update([m for m in prep_models if m])
            except:
                pass
    
    # Get from assembly machines
    for config in ASSEMBLY_CONFIGS:
        try:
            internal_models = config['prep_model'].objects.values_list('model_name_internal', flat=True).distinct()
            external_models = config['prep_model'].objects.values_list('model_name_external', flat=True).distinct()
            housing_models = config['prep_model'].objects.values_list('model_name_housing', flat=True).distinct()
            model_names.update([m for m in internal_models if m])
            model_names.update([m for m in external_models if m])
            model_names.update([m for m in housing_models if m])
        except:
            pass
    
    # Get from OP80
    try:
        op80_internal = OP80_CONFIG['prep_model'].objects.values_list('model_name_internal', flat=True).distinct()
        op80_external = OP80_CONFIG['prep_model'].objects.values_list('model_name_external', flat=True).distinct()
        model_names.update([m for m in op80_internal if m])
        model_names.update([m for m in op80_external if m])
    except:
        pass
    
    # Remove None and 'N/A'
    model_names.discard(None)
    model_names.discard('N/A')
    model_names.discard('')
    
    return sorted(list(model_names))


def get_machine_list_monitoring():
    """Get list of all machines for dropdown"""
    machines = []
    
    for config in MACHINE_CONFIGS:
        machines.append({
            'id': config['name'].lower().replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_'),
            'name': config['display_name'],
            'type': config['type']
        })
    
    for config in ASSEMBLY_CONFIGS:
        machines.append({
            'id': config['name'].lower().replace(' ', '_'),
            'name': config['display_name'],
            'type': 'assembly'
        })
    
    machines.append({
        'id': 'op80_leak_test',
        'name': OP80_CONFIG['display_name'],
        'type': 'op80'
    })
    
    return machines


def parse_time_filter(time_filter):
    """Convert time filter string to start_date and end_date"""
    end_dt = timezone.now()
    
    if time_filter == '15min':
        start_dt = end_dt - timedelta(minutes=15)
    elif time_filter == '30min':
        start_dt = end_dt - timedelta(minutes=30)
    elif time_filter == '1hour':
        start_dt = end_dt - timedelta(hours=1)
    elif time_filter == '2hour':
        start_dt = end_dt - timedelta(hours=2)
    elif time_filter == '4hour':
        start_dt = end_dt - timedelta(hours=4)
    elif time_filter == '6hour':
        start_dt = end_dt - timedelta(hours=6)
    elif time_filter == '12hour':
        start_dt = end_dt - timedelta(hours=12)
    elif time_filter == '24hour':
        start_dt = end_dt - timedelta(hours=24)
    elif time_filter.endswith('hour'):
        # Custom hours (e.g., "48hour", "72hour")
        try:
            hours = int(time_filter.replace('hour', ''))
            start_dt = end_dt - timedelta(hours=hours)
        except:
            # Default to last 1 hour if parsing fails
            start_dt = end_dt - timedelta(hours=1)
    else:
        # Default to last 1 hour
        start_dt = end_dt - timedelta(hours=1)
    
    return start_dt, end_dt


def search_monitoring_data(filters):
    """
    Search across all machines based on filters for monitoring
    Returns list of records with ALL details including gauge values, multiple QR codes, etc.
    """
    results = []
    
    # Extract filters
    qr_code = filters.get('qr_code', '').strip()
    model_name = filters.get('model_name', '')
    machine_id = filters.get('machine', '')
    time_filter = filters.get('time_filter', '1hour')
    custom_start = filters.get('start_date', '')
    custom_end = filters.get('end_date', '')
    status_filter = filters.get('status', '')
    
    # Parse dates based on time filter or custom range
    if custom_start and custom_end:
        try:
            start_dt = timezone.make_aware(datetime.strptime(custom_start, '%Y-%m-%d'))
        except:
            start_dt = timezone.now() - timedelta(hours=1)
        
        try:
            end_dt = timezone.make_aware(datetime.strptime(custom_end, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except:
            end_dt = timezone.now()
    else:
        start_dt, end_dt = parse_time_filter(time_filter)
    
    # Determine which machines to search
    configs_to_search = []
    if machine_id and machine_id != 'all':
        for config in MACHINE_CONFIGS:
            config_id = config['name'].lower().replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')
            if config_id == machine_id:
                configs_to_search.append(config)
                break
        
        if not configs_to_search:
            for config in ASSEMBLY_CONFIGS:
                config_id = config['name'].lower().replace(' ', '_')
                if config_id == machine_id:
                    configs_to_search.append(config)
                    break
        
        if not configs_to_search and machine_id == 'op80_leak_test':
            configs_to_search.append(OP80_CONFIG)
    else:
        configs_to_search = MACHINE_CONFIGS + ASSEMBLY_CONFIGS + [OP80_CONFIG]
    
    # Search each machine
    for config in configs_to_search:
        machine_name = config['name']
        display_name = config.get('display_name', machine_name)
        machine_type = config.get('type', 'standard')
        operation = config.get('operation', None)
        is_assembly = 'OP40' in machine_name
        is_washing = machine_type == 'washing'
        is_gauge = machine_type == 'gauge'
        is_painting = 'Painting' in machine_name
        is_lubrication = 'Lubrication' in machine_name
        is_op80 = 'Oring_leak' in machine_name
        
        # Handle washing machines separately
        if is_washing:
            if operation == 'load' and config.get('prep_model'):
                results.extend(search_washing_load_monitoring(config, qr_code, model_name, start_dt, end_dt, status_filter))
            elif operation == 'unload' and config.get('post_model'):
                results.extend(search_washing_unload_monitoring(config, qr_code, model_name, start_dt, end_dt, status_filter))
            continue
        
        if not config.get('prep_model'):
            continue
        
        # Build query for preprocessing
        query = Q()
        
        # QR Code filter
        if qr_code:
            if is_assembly:
                query &= (Q(qr_data_internal__icontains=qr_code) | 
                         Q(qr_data_external__icontains=qr_code) | 
                         Q(qr_data_housing__icontains=qr_code))
            elif is_painting:
                query &= (Q(qr_data_housing__icontains=qr_code) | Q(qr_data_piston__icontains=qr_code))
            elif is_lubrication:
                query &= (Q(qr_data_piston__icontains=qr_code) | Q(qr_data_housing__icontains=qr_code))
            elif is_op80:
                query &= (Q(qr_data_piston__icontains=qr_code) | Q(qr_data_housing__icontains=qr_code))
            else:
                query &= Q(qr_data__icontains=qr_code)
        
        # Model name filter
        if model_name and model_name != 'all':
            if is_assembly:
                query &= (Q(model_name_internal=model_name) | 
                         Q(model_name_external=model_name) | 
                         Q(model_name_housing=model_name))
            elif is_painting:
                query &= (Q(model_name_housing=model_name) | Q(model_name_piston=model_name))
            elif is_lubrication:
                query &= (Q(model_name_piston=model_name) | Q(model_name_housing=model_name))
            elif is_op80:
                query &= (Q(model_name_internal=model_name) | Q(model_name_external=model_name))
            else:
                query &= Q(model_name=model_name)
        
        # Get preprocessing records
        try:
            prep_records = config['prep_model'].objects.filter(query)[:1000]
        except Exception as e:
            print(f"Error querying {machine_name}: {e}")
            continue
        
        # Process each record
        for prep in prep_records:
            # Get timestamp
            if is_assembly:
                timestamp = prep.timestamp_internal
            else:
                timestamp = prep.timestamp
            
            # Date filter
            if start_dt or end_dt:
                dt = parse_timestamp_to_datetime(timestamp)
                if dt:
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt)
                    
                    if start_dt and dt < start_dt:
                        continue
                    if end_dt and dt > end_dt:
                        continue
            
            # Initialize record with common fields
            record = {
                'prep_id': prep.id,
                'machine_name': machine_name,
                'display_name': display_name,
                'machine_type': machine_type,
            }
            
            # Get status and details based on machine type
            if is_assembly:
                # ASSEMBLY MACHINES
                qr_internal = prep.qr_data_internal
                qr_external = prep.qr_data_external or '-'
                qr_housing = prep.qr_data_housing or '-'
                status = prep.status if prep.qr_data_external and prep.qr_data_housing else 'Pending'
                
                record.update({
                    'qr_code': qr_internal,
                    'qr_internal': qr_internal,
                    'qr_external': qr_external,
                    'qr_housing': qr_housing,
                    'model_name': getattr(prep, 'model_name_internal', 'N/A'),
                    'model_name_internal': getattr(prep, 'model_name_internal', 'N/A'),
                    'model_name_external': getattr(prep, 'model_name_external', 'N/A'),
                    'model_name_housing': getattr(prep, 'model_name_housing', 'N/A'),
                    'previous_machine_internal_status': getattr(prep, 'previous_machine_internal_status', '-'),
                    'previous_machine_housing_status': getattr(prep, 'previous_machine_housing_status', '-'),
                    'status': status,
                    'post_id': prep.id if status != 'Pending' else None,
                })
                
            elif is_painting:
                # PAINTING MACHINE
                qr_housing = prep.qr_data_housing
                qr_piston = prep.qr_data_piston
                post = config['post_model'].objects.filter(qr_data_housing=qr_housing).first()
                status = post.status if post else 'Pending'
                
                record.update({
                    'qr_code': qr_housing,
                    'qr_housing': qr_housing,
                    'qr_piston': qr_piston,
                    'model_name': getattr(prep, 'model_name_housing', 'N/A'),
                    'model_name_housing': getattr(prep, 'model_name_housing', 'N/A'),
                    'model_name_piston': getattr(prep, 'model_name_piston', 'N/A'),
                    'previous_machine_status': getattr(prep, 'previous_machine_status', '-'),
                    'pre_status': getattr(prep, 'pre_status', '-'),
                    'status': status,
                    'post_id': post.id if post else None,
                })
                
            elif is_lubrication:
                # LUBRICATION MACHINE
                qr_piston = prep.qr_data_piston
                qr_housing = prep.qr_data_housing
                post = config['post_model'].objects.filter(qr_data_piston=qr_piston).first()
                status = post.status if post else 'Pending'
                
                record.update({
                    'qr_code': qr_piston,
                    'qr_piston': qr_piston,
                    'qr_housing': qr_housing,
                    'model_name': getattr(prep, 'model_name_piston', 'N/A'),
                    'model_name_piston': getattr(prep, 'model_name_piston', 'N/A'),
                    'model_name_housing': getattr(prep, 'model_name_housing', 'N/A'),
                    'previous_machine_status': getattr(prep, 'previous_machine_status', '-'),
                    'status': status,
                    'post_id': post.id if post else None,
                })
                
            elif is_op80:
                # OP80 LEAK TEST
                qr_piston = prep.qr_data_piston
                qr_housing = prep.qr_data_housing
                
                # Try multiple matching strategies
                post = config['post_model'].objects.filter(qr_data_housing=qr_housing).first()
                if not post:
                    post = config['post_model'].objects.filter(qr_data_housing_new=qr_housing).first()
                if not post:
                    post = config['post_model'].objects.filter(qr_data_housing=qr_piston).first()
                
                status = post.status if post else 'Pending'
                
                record.update({
                    'qr_code': qr_piston,
                    'qr_piston': qr_piston,
                    'qr_housing': qr_housing,
                    'qr_housing_new': post.qr_data_housing_new if post else '-',
                    'model_name': getattr(prep, 'model_name_internal', 'N/A'),
                    'model_name_internal': getattr(prep, 'model_name_internal', 'N/A'),
                    'model_name_external': getattr(prep, 'model_name_external', 'N/A'),
                    'previous_machine_status': getattr(prep, 'previous_machine_status', '-'),
                    'match_status': post.match_status if post else '-',
                    'status': status,
                    'post_id': post.id if post else None,
                })
                
            else:
                # STANDARD MACHINES (CNC, Gauge, Honing, etc.)
                qr_value = prep.qr_data
                post = config['post_model'].objects.filter(qr_data=qr_value).first() if config['post_model'] else None
                status = post.status if post else 'Pending'
                
                record.update({
                    'qr_code': qr_value,
                    'model_name': getattr(prep, 'model_name', 'N/A'),
                    'previous_machine_status': getattr(prep, 'previous_machine_status', '-'),
                    'status': status,
                    'post_id': post.id if post else None,
                })
                
                # Add gauge values if this is a gauge machine
                if is_gauge and post:
                    gauge_values = {}
                    for i in range(1, 7):
                        val = getattr(post, f'value{i}', None)
                        if val is not None:
                            gauge_values[f'value{i}'] = val
                    record['gauge_values'] = gauge_values
            
            # Status filter
            if status_filter and status_filter != 'all' and status != status_filter:
                continue
            
            # Format timestamp
            if hasattr(timestamp, 'isoformat'):
                timestamp_str = timestamp.isoformat()
            else:
                timestamp_str = str(timestamp)
            
            record['timestamp'] = timestamp_str
            results.append(record)
    
    # Sort by timestamp (newest first)
    results.sort(key=lambda x: x['timestamp'], reverse=True)
    
    return results


def search_washing_load_monitoring(config, qr_code, model_name, start_dt, end_dt, status_filter):
    """Search washing machine loading stage"""
    results = []
    prep_model = config['prep_model']
    
    query = Q()
    if qr_code:
        query &= Q(qr_data__icontains=qr_code)
    if model_name and model_name != 'all':
        query &= Q(model_name=model_name)
    
    try:
        prep_records = prep_model.objects.filter(query)[:1000]
    except Exception as e:
        return results
    
    for prep in prep_records:
        timestamp = prep.timestamp
        
        if start_dt or end_dt:
            dt = parse_timestamp_to_datetime(timestamp)
            if dt:
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
                if start_dt and dt < start_dt:
                    continue
                if end_dt and dt > end_dt:
                    continue
        
        status = getattr(prep, 'status', 'OK')
        
        if status_filter and status_filter != 'all' and status != status_filter:
            continue
        
        if hasattr(timestamp, 'isoformat'):
            timestamp_str = timestamp.isoformat()
        else:
            timestamp_str = str(timestamp)
        
        record = {
            'prep_id': prep.id,
            'post_id': None,
            'machine_name': config['name'],
            'display_name': config['display_name'],
            'machine_type': 'washing',
            'operation': 'load',
            'stage': 'Loading (Preprocessing)',
            'qr_code': prep.qr_data,
            'model_name': getattr(prep, 'model_name', 'N/A'),
            'timestamp': timestamp_str,
            'status': status,
            'previous_machine_status': getattr(prep, 'previous_machine_status', '-'),
        }
        
        results.append(record)
    
    return results


def search_washing_unload_monitoring(config, qr_code, model_name, start_dt, end_dt, status_filter):
    """Search washing machine unloading stage"""
    results = []
    post_model = config['post_model']
    
    query = Q()
    if qr_code:
        query &= Q(qr_data__icontains=qr_code)
    if model_name and model_name != 'all':
        query &= Q(model_name=model_name)
    
    try:
        post_records = post_model.objects.filter(query)[:1000]
    except Exception as e:
        return results
    
    for post in post_records:
        timestamp = post.timestamp
        
        if start_dt or end_dt:
            dt = parse_timestamp_to_datetime(timestamp)
            if dt:
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
                if start_dt and dt < start_dt:
                    continue
                if end_dt and dt > end_dt:
                    continue
        
        status = getattr(post, 'status', 'OK')
        
        if status_filter and status_filter != 'all' and status != status_filter:
            continue
        
        if hasattr(timestamp, 'isoformat'):
            timestamp_str = timestamp.isoformat()
        else:
            timestamp_str = str(timestamp)
        
        record = {
            'prep_id': None,
            'post_id': post.id,
            'machine_name': config['name'],
            'display_name': config['display_name'],
            'machine_type': 'washing',
            'operation': 'unload',
            'stage': 'Unloading (Postprocessing)',
            'qr_code': post.qr_data,
            'model_name': getattr(post, 'model_name', 'N/A'),
            'timestamp': timestamp_str,
            'status': status,
            'previous_machine_status': getattr(post, 'previous_machine_status', '-'),
        }
        
        results.append(record)
    
    return results


def monitoring_page(request):
    """Main monitoring page view"""
    model_names = get_all_model_names_monitoring()
    machines = get_machine_list_monitoring()
    
    context = {
        'model_names': model_names,
        'machines': machines,
    }
    
    return render(request, 'rework/monitoring.html', context)


@csrf_exempt
def monitoring_search_api(request):
    """API endpoint for searching records"""
    if request.method == 'GET':
        filters = {
            'qr_code': request.GET.get('qr_code', ''),
            'model_name': request.GET.get('model_name', ''),
            'machine': request.GET.get('machine', ''),
            'time_filter': request.GET.get('time_filter', '1hour'),
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'status': request.GET.get('status', ''),
        }
        
        results = search_monitoring_data(filters)
        
        return JsonResponse({
            'success': True,
            'count': len(results),
            'results': results
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)


@csrf_exempt
def monitoring_export_excel(request):
    """Export monitoring data to Excel with ALL details and styling"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    # Get the same filters used for search
    filters = {
        'qr_code': request.GET.get('qr_code', ''),
        'model_name': request.GET.get('model_name', ''),
        'machine': request.GET.get('machine', ''),
        'time_filter': request.GET.get('time_filter', '1hour'),
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'status': request.GET.get('status', ''),
    }
    
    # Get data using same search function
    results = search_monitoring_data(filters)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitoring Data"
    
    # Define styles
    header_fill = PatternFill(start_color="FF7755", end_color="FF7755", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Status-based fills
    status_fills = {
        'OK': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'NG': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'Pending': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }
    
    # Comprehensive headers - include ALL possible fields
    headers = [
        'Prep ID', 'Post ID', 'Machine', 'Machine Type', 'QR Code', 
        'Date', 'Time', 'Status', 'Model Name', 'Previous Status',
        # Assembly specific
        'QR Internal', 'QR External', 'QR Housing',
        'Model Internal', 'Model External', 'Model Housing',
        'Prev Internal Status', 'Prev Housing Status',
        # Painting specific
        'QR Piston', 'Model Piston', 'Model Housing (Paint)', 'Pre-Status',
        # Lubrication specific  
        'QR Piston (Lub)', 'QR Housing (Lub)', 'Model Piston (Lub)', 'Model Housing (Lub)',
        # OP80 specific
        'QR Piston (OP80)', 'QR Housing (OP80)', 'QR Housing New', 'Model Internal (OP80)', 
        'Model External (OP80)', 'Match Status',
        # Gauge specific
        'Gauge Value 1', 'Gauge Value 2', 'Gauge Value 3', 
        'Gauge Value 4', 'Gauge Value 5', 'Gauge Value 6',
        # Washing specific
        'Washing Stage', 'Washing Prev Status'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
        ws.row_dimensions[1].height = 30
    
    # Data rows
    for row_num, record in enumerate(results, 2):
        col_num = 1
        
        # Common fields
        ws.cell(row=row_num, column=col_num, value=record.get('prep_id') or '-')
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=record.get('post_id') or '-')
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=record.get('display_name', 'N/A'))
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=record.get('machine_type', 'N/A'))
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=record.get('qr_code', '-'))
        col_num += 1
        
        # Date and Time
        timestamp_str = record.get('timestamp', '')
        try:
            dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            ws.cell(row=row_num, column=col_num, value=dt.strftime('%Y-%m-%d'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=dt.strftime('%H:%M:%S'))
            col_num += 1
        except:
            ws.cell(row=row_num, column=col_num, value=timestamp_str)
            col_num += 1
            ws.cell(row=row_num, column=col_num, value='')
            col_num += 1
        
        # Status with color
        status = record.get('status', 'Pending')
        status_cell = ws.cell(row=row_num, column=col_num, value=status)
        if status in status_fills:
            status_cell.fill = status_fills[status]
        status_cell.font = Font(bold=True)
        col_num += 1
        
        # Model Name and Previous Status
        ws.cell(row=row_num, column=col_num, value=record.get('model_name', 'N/A'))
        col_num += 1
        ws.cell(row=row_num, column=col_num, value=record.get('previous_machine_status', '-'))
        col_num += 1
        
        # Assembly specific fields
        machine_type = record.get('machine_type', '')
        if machine_type == 'assembly':
            ws.cell(row=row_num, column=col_num, value=record.get('qr_internal', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('qr_external', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('qr_housing', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_internal', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_external', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_housing', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('previous_machine_internal_status', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('previous_machine_housing_status', '-'))
            col_num += 1
        else:
            for _ in range(8):
                ws.cell(row=row_num, column=col_num, value='-')
                col_num += 1
        
        # Painting specific fields
        if record.get('machine_name', '').find('Painting') != -1:
            ws.cell(row=row_num, column=col_num, value=record.get('qr_piston', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_piston', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_housing', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('pre_status', '-'))
            col_num += 1
        else:
            for _ in range(4):
                ws.cell(row=row_num, column=col_num, value='-')
                col_num += 1
        
        # Lubrication specific fields
        if record.get('machine_name', '').find('Lubrication') != -1:
            ws.cell(row=row_num, column=col_num, value=record.get('qr_piston', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('qr_housing', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_piston', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_housing', 'N/A'))
            col_num += 1
        else:
            for _ in range(4):
                ws.cell(row=row_num, column=col_num, value='-')
                col_num += 1
        
        # OP80 specific fields
        if machine_type == 'op80':
            ws.cell(row=row_num, column=col_num, value=record.get('qr_piston', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('qr_housing', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('qr_housing_new', '-'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_internal', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('model_name_external', 'N/A'))
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('match_status', '-'))
            col_num += 1
        else:
            for _ in range(6):
                ws.cell(row=row_num, column=col_num, value='-')
                col_num += 1
        
        # Gauge values
        gauge_values = record.get('gauge_values', {})
        for i in range(1, 7):
            val = gauge_values.get(f'value{i}', '-')
            ws.cell(row=row_num, column=col_num, value=val)
            col_num += 1
        
        # Washing specific fields
        if machine_type == 'washing':
            operation = record.get('operation', '')
            stage = 'Loading' if operation == 'load' else 'Unloading' if operation == 'unload' else '-'
            ws.cell(row=row_num, column=col_num, value=stage)
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=record.get('previous_machine_status', '-'))
            col_num += 1
        else:
            for _ in range(2):
                ws.cell(row=row_num, column=col_num, value='-')
                col_num += 1
        
        # Apply borders to all cells in row
        for c in range(1, col_num):
            ws.cell(row=row_num, column=c).border = border_style
            ws.cell(row=row_num, column=c).alignment = Alignment(vertical="center")
    
    # Auto-fit column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = len(headers[col_num - 1])
        for cell in ws[column_letter][1:]:  # Skip header
            try:
                if cell.value and str(cell.value) != '-':
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row and first 5 columns
    ws.freeze_panes = 'F2'
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="monitoring_detailed_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response



@csrf_exempt
def monitoring_chart_data_api(request):
    """API endpoint for chart data"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    # Get the same filters
    filters = {
        'qr_code': request.GET.get('qr_code', ''),
        'model_name': request.GET.get('model_name', ''),
        'machine': request.GET.get('machine', ''),
        'time_filter': request.GET.get('time_filter', '1hour'),
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
        'status': request.GET.get('status', ''),
    }
    
    # Get data
    results = search_monitoring_data(filters)
    
    # Prepare chart data
    from collections import defaultdict
    
    # Time-based aggregation (per 5-minute intervals)
    time_status_data = defaultdict(lambda: {'OK': 0, 'NG': 0, 'Pending': 0})
    machine_status_data = defaultdict(lambda: {'OK': 0, 'NG': 0, 'Pending': 0})
    model_status_data = defaultdict(lambda: {'OK': 0, 'NG': 0, 'Pending': 0})
    
    for record in results:
        # Time aggregation
        timestamp_str = record.get('timestamp', '')
        try:
            dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
            # Round to 5-minute intervals
            time_key = dt.replace(minute=(dt.minute // 5) * 5, second=0, microsecond=0).strftime('%H:%M')
        except:
            time_key = 'Unknown'
        
        status = record.get('status', 'Pending')
        time_status_data[time_key][status] += 1
        
        # Machine aggregation
        machine_name = record.get('display_name', 'Unknown')
        machine_status_data[machine_name][status] += 1
        
        # Model aggregation
        model_name = record.get('model_name', 'N/A')
        if model_name != 'N/A':
            model_status_data[model_name][status] += 1
    
    # Format for charts
    sorted_times = sorted(time_status_data.keys())
    
    chart_data = {
        'time_series': {
            'labels': sorted_times,
            'ok': [time_status_data[t]['OK'] for t in sorted_times],
            'ng': [time_status_data[t]['NG'] for t in sorted_times],
            'pending': [time_status_data[t]['Pending'] for t in sorted_times],
        },
        'machine_breakdown': {
            'machines': list(machine_status_data.keys()),
            'ok': [machine_status_data[m]['OK'] for m in machine_status_data.keys()],
            'ng': [machine_status_data[m]['NG'] for m in machine_status_data.keys()],
            'pending': [machine_status_data[m]['Pending'] for m in machine_status_data.keys()],
        },
        'model_breakdown': {
            'models': list(model_status_data.keys()),
            'ok': [model_status_data[m]['OK'] for m in model_status_data.keys()],
            'ng': [model_status_data[m]['NG'] for m in model_status_data.keys()],
            'pending': [model_status_data[m]['Pending'] for m in model_status_data.keys()],
        },
        'summary': {
            'total': len(results),
            'ok': sum(r['status'] == 'OK' for r in results),
            'ng': sum(r['status'] == 'NG' for r in results),
            'pending': sum(r['status'] == 'Pending' for r in results),
        }
    }
    
    return JsonResponse({
        'success': True,
        'data': chart_data
    })